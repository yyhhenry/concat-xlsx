import argparse
import os

from openpyxl import Workbook, load_workbook

parser = argparse.ArgumentParser(description="合并多个xlsx文件到一个文件中")
parser.add_argument(
    "input_dir",
    type=str,
    default="./",
    help="指定文件夹路径",
)
parser.add_argument(
    "-o",
    "--output",
    dest="output",
    type=str,
    default="output.xlsx",
    help="目标文件名",
)
parser.add_argument(
    "--without-header",
    dest="without_header",
    action="store_true",
    help="不把第一行看作表头",
)
args = parser.parse_args()
input_dir = str(args.input_dir)
output_file = str(args.output)
without_header = bool(args.without_header)

xlsx_files = [f for f in os.listdir(input_dir) if f.endswith(".xlsx")]

target_wb = Workbook()
target_ws = target_wb[target_wb.sheetnames[0]]

offset = 0

for file in xlsx_files:
    print(f"正在处理文件：{file}")
    wb = load_workbook(os.path.join(input_dir, file), read_only=True)
    ws = wb.active
    assert ws is not None, "没有找到活动工作表！"

    start = 1 if offset == 0 or without_header else 2  # 跳过表头，第一个或无表头除外

    for row in ws.iter_rows(min_row=start, values_only=False):
        for col, cell in enumerate(row):
            new_cell = target_ws.cell(row=offset + 1, column=col + 1)
            new_cell.value = cell.value

            if cell.value is not None and cell.has_style:
                new_cell.number_format = cell.number_format
                new_cell.alignment += cell.alignment

        offset += 1


target_wb.save(output_file)

print("合并完成！")
